"""
Order Import API - Handle Excel file uploads and import orders into database
"""

from flask import Blueprint, request, jsonify
from datetime import datetime
import json
from openpyxl import load_workbook
from io import BytesIO

from backend.utils.auth import auth_required, require_permission
from backend.utils.logger import logger
from backend.utils.response import success_response, error_response
from backend.config.database import get_db_connection

order_import_bp = Blueprint('order_import', __name__, url_prefix='/api/orders/import')


@order_import_bp.route('/upload', methods=['POST'])
@auth_required
@require_permission('planning', 'write')
def upload_order_file():
    """
    Upload and preview Excel file before importing
    """
    try:
        if 'file' not in request.files:
            return error_response('No file provided', 400)
        
        file = request.files['file']
        if file.filename == '':
            return error_response('No file selected', 400)
        
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return error_response('Only Excel (.xlsx, .xls) and CSV files are supported', 400)
        
        # Read the file
        file_content = file.read()
        
        # Parse based on file type
        if file.filename.endswith('.csv'):
            rows = parse_csv(file_content)
        else:
            rows = parse_excel(file_content)
        
        if not rows:
            return error_response('File is empty or could not be parsed', 400)
        
        # Parse header and data rows
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        
        # Map Excel columns to database fields
        mapped_data = []
        errors = []
        
        for idx, row in enumerate(data_rows, start=2):  # Start at row 2 (after header)
            try:
                mapped_row = map_excel_row_to_order(row, headers)
                if mapped_row:
                    mapped_row['row_number'] = idx
                    mapped_data.append(mapped_row)
            except Exception as e:
                errors.append({'row': idx, 'error': str(e)})
        
        return success_response({
            'preview': mapped_data[:10],  # Show first 10 rows for preview
            'total_rows': len(mapped_data),
            'errors': errors,
            'headers': headers,
            'file_name': file.filename
        })
    
    except Exception as e:
        logger.error(f"Error uploading order file: {str(e)}")
        return error_response(f'Error processing file: {str(e)}', 500)


@order_import_bp.route('/import', methods=['POST'])
@auth_required
@require_permission('planning', 'write')
def import_orders():
    """
    Import orders from Excel file into database
    """
    try:
        data = request.get_json()
        
        if 'file_content' not in data or 'headers' not in data:
            return error_response('Missing required data', 400)
        
        file_content = data['file_content']
        headers = data['headers']
        
        # Parse file content (base64 encoded)
        import base64
        file_bytes = base64.b64decode(file_content)
        rows = parse_excel(file_bytes)
        
        if not rows or len(rows) < 2:
            return error_response('No data rows found in file', 400)
        
        data_rows = rows[1:]
        db = get_db_connection()
        cursor = db.cursor()
        
        imported_count = 0
        error_details = []
        
        try:
            for idx, row in enumerate(data_rows, start=2):
                try:
                    order_data = map_excel_row_to_order(row, headers)
                    if order_data:
                        result = create_order_from_excel(cursor, order_data)
                        if result['success']:
                            imported_count += 1
                        else:
                            error_details.append({
                                'row': idx,
                                'error': result['error'],
                                'reference': order_data.get('reference_number', 'Unknown')
                            })
                except Exception as e:
                    error_details.append({
                        'row': idx,
                        'error': str(e)
                    })
            
            db.commit()
            
            return success_response({
                'imported_count': imported_count,
                'total_rows': len(data_rows),
                'errors': error_details,
                'message': f'Successfully imported {imported_count} orders'
            })
        
        except Exception as e:
            db.rollback()
            raise e
        
        finally:
            cursor.close()
            db.close()
    
    except Exception as e:
        logger.error(f"Error importing orders: {str(e)}")
        return error_response(f'Error importing orders: {str(e)}', 500)


def parse_excel(file_content):
    """
    Parse Excel file and return rows
    """
    try:
        file_obj = BytesIO(file_content)
        workbook = load_workbook(file_obj)
        worksheet = workbook.active
        
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            # Convert tuple to list and handle None values
            row_data = [cell if cell is not None else '' for cell in row]
            rows.append(row_data)
        
        return rows
    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        raise


def parse_csv(file_content):
    """
    Parse CSV file and return rows
    """
    try:
        import csv
        file_str = file_content.decode('utf-8')
        reader = csv.reader(file_str.split('\n'))
        rows = list(reader)
        return [row for row in rows if any(row)]  # Filter empty rows
    except Exception as e:
        logger.error(f"Error parsing CSV file: {str(e)}")
        raise


def map_excel_row_to_order(row, headers):
    """
    Map Excel row to order data structure
    """
    # Create a dictionary mapping headers to values
    row_dict = {}
    for i, header in enumerate(headers):
        if i < len(row):
            row_dict[header.lower().strip()] = row[i]
    
    # Extract relevant fields
    order_data = {
        'priority': map_priority(row_dict.get('priority', '')),
        'reference_number': row_dict.get('reference number', ''),
        'pool': row_dict.get('pool', ''),
        'customer_name': row_dict.get('customer name', ''),
        'status': map_status(row_dict.get('status', '')),
        'quantity': parse_number(row_dict.get('quantity', 0)),
        'production_end': parse_date(row_dict.get('production end', '')),
        'delivery_date': parse_date(row_dict.get('delivery', '')),
        'artwork_approved': row_dict.get('artwork approved', '').lower() in ['yes', 'true', '1'],
        'height': parse_number(row_dict.get('height', 0)),
        'width': parse_number(row_dict.get('with', 0)),  # Note: Excel has "With" instead of "Width"
        'depth': parse_number(row_dict.get('depth', 0)),
        'stitch_count_color': row_dict.get('stitch count color', ''),
        'on_hold': row_dict.get('on hold', '').lower() in ['yes', 'true', '1'],
        'notes': row_dict.get('moved out notes', ''),
        'estimated_amount': parse_decimal(row_dict.get('estimated amount', 0)),
        'colour': row_dict.get('colour', ''),
        'machine_number': row_dict.get('machine number', ''),
        'run_time_hours': parse_number(row_dict.get('run time (hours)', 0)),
        'run_time_minutes': parse_number(row_dict.get('run time (minutes)', 0)),
        'kevro_status': row_dict.get('kevro status', ''),
        'production_code': row_dict.get('production', ''),
        'item_number': row_dict.get('item number', ''),
        'sales_district': row_dict.get('sales district', ''),
        'product_name': row_dict.get('name', ''),
        'production_start': parse_date(row_dict.get('production start', '')),
        'mode_of_delivery': row_dict.get('mode of delivery', ''),
        'created_date': parse_datetime(row_dict.get('created date and time', '')),
        'transfer_number': row_dict.get('transfer number', ''),
        'density': parse_number(row_dict.get('density', 0)),
        'positions': parse_number(row_dict.get('positions', 0)),
    }
    
    # Validate required fields
    if not order_data['reference_number']:
        raise ValueError('Reference number is required')
    
    if not order_data['customer_name']:
        raise ValueError('Customer name is required')
    
    return order_data


def create_order_from_excel(cursor, order_data):
    """
    Create order in database from Excel data
    """
    try:
        # Get or create product
        product_id = get_or_create_product(cursor, order_data)
        
        # Check if order already exists
        cursor.execute(
            'SELECT id FROM orders WHERE order_number = %s',
            (order_data['reference_number'],)
        )
        existing = cursor.fetchone()
        
        if existing:
            return {'success': False, 'error': 'Order already exists with this reference number'}
        
        # Prepare order data for insertion
        order_insert_data = {
            'order_number': order_data['reference_number'],
            'customer_name': order_data['customer_name'],
            'product_id': product_id,
            'quantity': order_data['quantity'],
            'order_value': order_data['estimated_amount'],
            'start_date': order_data['production_start'],
            'end_date': order_data['production_end'],
            'priority': order_data['priority'],
            'status': order_data['status'],
            'hold_reason': 'On hold' if order_data['on_hold'] else None,
            'notes': order_data['notes'],
            'config': json.dumps({
                'pool': order_data['pool'],
                'artwork_approved': order_data['artwork_approved'],
                'height': order_data['height'],
                'width': order_data['width'],
                'depth': order_data['depth'],
                'stitch_count_color': order_data['stitch_count_color'],
                'colour': order_data['colour'],
                'machine_number': order_data['machine_number'],
                'run_time_hours': order_data['run_time_hours'],
                'run_time_minutes': order_data['run_time_minutes'],
                'kevro_status': order_data['kevro_status'],
                'production_code': order_data['production_code'],
                'item_number': order_data['item_number'],
                'sales_district': order_data['sales_district'],
                'production_start': str(order_data['production_start']),
                'mode_of_delivery': order_data['mode_of_delivery'],
                'transfer_number': order_data['transfer_number'],
                'density': order_data['density'],
                'positions': order_data['positions'],
            })
        }
        
        # Insert order
        columns = ', '.join(order_insert_data.keys())
        placeholders = ', '.join(['%s'] * len(order_insert_data))
        values = tuple(order_insert_data.values())
        
        cursor.execute(
            f'INSERT INTO orders ({columns}) VALUES ({placeholders})',
            values
        )
        
        return {'success': True, 'order_id': cursor.lastrowid}
    
    except Exception as e:
        logger.error(f"Error creating order: {str(e)}")
        return {'success': False, 'error': str(e)}


def get_or_create_product(cursor, order_data):
    """
    Get existing product or create new one
    """
    product_code = order_data['product_name'][:50] if order_data['product_name'] else 'UNKNOWN'
    
    # Check if product exists
    cursor.execute(
        'SELECT id FROM products WHERE product_code = %s',
        (product_code,)
    )
    result = cursor.fetchone()
    
    if result:
        return result[0]
    
    # Create new product
    product_data = {
        'product_code': product_code,
        'product_name': order_data['product_name'],
        'category': order_data['pool'],
        'specifications': json.dumps({
            'height': order_data['height'],
            'width': order_data['width'],
            'depth': order_data['depth'],
            'colour': order_data['colour'],
            'density': order_data['density'],
        })
    }
    
    cursor.execute(
        'INSERT INTO products (product_code, product_name, category, specifications) VALUES (%s, %s, %s, %s)',
        (product_data['product_code'], product_data['product_name'], product_data['category'], product_data['specifications'])
    )
    
    return cursor.lastrowid


def map_priority(priority_str):
    """
    Map priority string to database value
    """
    priority_map = {
        '1': 'normal',
        '2': 'high',
        '3': 'urgent',
        'high': 'high',
        'urgent': 'urgent',
        'normal': 'normal',
        'low': 'low',
    }
    return priority_map.get(str(priority_str).lower().strip(), 'normal')


def map_status(status_str):
    """
    Map status string to database value
    """
    status_map = {
        'scheduled': 'scheduled',
        'in production': 'in_progress',
        'in production - picked': 'in_progress',
        'completed': 'completed',
        'on hold': 'on_hold',
        'cancelled': 'cancelled',
        'unscheduled': 'unscheduled',
    }
    return status_map.get(status_str.lower().strip(), 'unscheduled')


def parse_number(value):
    """
    Parse number value, handling various formats
    """
    if not value or value == '':
        return 0
    try:
        # Handle comma as decimal separator
        value_str = str(value).replace(',', '.')
        return float(value_str)
    except:
        return 0


def parse_decimal(value):
    """
    Parse decimal value for currency
    """
    return parse_number(value)


def parse_date(date_str):
    """
    Parse date string in various formats
    """
    if not date_str or date_str == '':
        return None
    
    date_formats = [
        '%Y/%m/%d',
        '%Y-%m-%d',
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%m/%d/%Y',
        '%m-%d-%Y',
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(str(date_str).strip(), fmt).date()
        except:
            continue
    
    return None


def parse_datetime(datetime_str):
    """
    Parse datetime string in various formats
    """
    if not datetime_str or datetime_str == '':
        return None
    
    datetime_formats = [
        '%Y/%m/%d %H:%M',
        '%Y-%m-%d %H:%M:%S',
        '%d/%m/%Y %H:%M',
        '%Y/%m/%d %H:%M:%S',
    ]
    
    for fmt in datetime_formats:
        try:
            return datetime.strptime(str(datetime_str).strip(), fmt)
        except:
            continue
    
    return None
